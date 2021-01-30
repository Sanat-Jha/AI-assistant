# Imports
import pyttsx3
import speech_recognition as sr
import wikipedia
import pyautogui
import openpyxl
import webbrowser
from datetime import datetime,date
import os
from PyDictionary import PyDictionary
import pywhatkit
import random

#starters
engine = pyttsx3.init('sapi5')
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[0].id)
db = openpyxl.load_workbook("database.xlsx")
people = db["People"]
app_web = db["appweb"]
latest_reminder = ""

# Variables
Mode = "OFF"
Owner = "Sanat"


# Lists
Greets = [
    "Yes sir",
    "Yes boss",
    "Hello sir",
    "Hi dude",
    "Ha bol",
    "Hi " + Owner,
    "Yes " + Owner + "Sir"
]

reminders = [
    ["10:25","Join the class"],
    ["8:55","Join the class"],
    ["12:55","Join the class"],
    ["1:25","Join the class"],
    ]
# Functions


def speak(audio):
    engine.say(audio)
    engine.runAndWait()


def listen():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        r.pause_threshold = 0.3
        r.phrase_threshold = 0.3
        r.non_speaking_duration = 0.3
        audio = r.listen(source)

    try:
        query = r.recognize_google(audio, language="en-in")
    except Exception:
        return ""
    return query


# def speak(audio):
#     print(audio)


# def listen():
#     query = input(">> ")
#     return query


def mousedown(x):
    position = pyautogui.position()
    pyautogui.moveTo(position[0], position[1]+x)


def mouseup(x):
    position = pyautogui.position()
    pyautogui.moveTo(position[0], position[1]-x)


def mouseleft(x):
    position = pyautogui.position()
    pyautogui.moveTo(position[0]-x, position[1])


def mouseright(x):
    position = pyautogui.position()
    pyautogui.moveTo(position[0]+x, position[1])

def getnumber(x):
    for word in x.split():
        if word.isdigit():
            return int(word)

def getpath(x):
    for i in range(2,app_web.max_row+1):
        if app_web.cell(column=1,row=i).value.lower() == x.lower():
            return app_web.cell(column=2,row=i).value
    else:
        return None

def isApp(x):
    for i in range(2,app_web.max_row+1):
        if app_web.cell(column=1,row=i).value.lower() == x.lower():
            if app_web.cell(column=3,row=i).value == "App":
                return True
            elif app_web.cell(column=3,row=i).value == "Web":
                return False
    else:
        return None

def cleanText(text):
    uselessWords = [
        " a ",
        " for ",
        " an ",
        " the ",
        " by ",
        " what ",
        " how ",
        " in ",
        " on ",
        " the ",
        " at ",
        " is ",
        " are "
    ]
    for word in uselessWords:
        text = text.replace(word,"")
    return text


if __name__ == "__main__":
    now = datetime.now()
    current_time = now.strftime("%H:%M")
    if latest_reminder != current_time:
        for reminder in reminders:
            if reminder[0] == current_time:
                speak(reminder[1])
                latest_reminder = current_time
    while True:
        query = listen().lower()
        query = cleanText(query)
        print(query)
        if "robo" in query:
            query = query.replace("robo", "")
            speak(Greets[random.randint(0,len(Greets)-1)])
            Mode = "ON"
        if Mode == "ON":
            if "wikipedia" in query:
                try:
                    query = query.replace("wikipedia", "")
                    answer = wikipedia.summary(query, 2)
                    speak(answer)
                except Exception:
                    speak("Some error occured while searching wikipedia. Please try again.")

            if "mouse up" in query:
                query = query.replace("mouse", "")
                query = query.replace("up", "")
                mouseup(getnumber(query))
            if "mouse down" in query:
                query = query.replace("mouse", "")
                query = query.replace("down", "")
                mousedown(getnumber(query))
            if "mouse left" in query:
                query = query.replace("mouse", "")
                query = query.replace("left", "")
                mouseleft(getnumber(query))
            if "mouse right" in query:
                query = query.replace("mouse", "")
                query = query.replace("right", "")
                mouseright(getnumber(query))
            if "click" in query:
                query = query.replace("click", "")
                pyautogui.click()

            if "open" in query:
                try:
                    query = query.replace("open", "")
                    query = query.replace(" ", "")
                    if isApp(query):
                        path = getpath(query)
                        print(path)
                        os.startfile(path)
                    
                    elif isApp(query) == None:
                        speak("No such app or website found.")

                    else:
                        path = getpath(query)
                        print(path)
                        webbrowser.get("C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s").open(path)
                except:
                    speak("some issue occured ")

            if "search youtube" in query:
                query = query.replace("search youtube", "")
                webbrowser.get("C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s").open("https://www.youtube.com/results?search_query="+query)
            
            if "time" in query:
                query = query.replace("time", "")
                now = datetime.now()
                current_time = now.strftime("%H:%M")
                speak("Its "+current_time)

            if "date" in query :
                query = query.replace("date", "")
                query = query.replace("today", "")
                today = date.today()
                speak("Today's date is" + str(today))
            
            if "meaning of" in query:
                query = query.replace("meaning of", "")
                MyDict = PyDictionary(query)
                speak(MyDict.getMeanings())
            if "google" in query:
                query = query.replace("google", "")
                pywhatkit.search(query)
            
            if "screenshot" in query:
                query = query.replace("screenshot", "")
                img = pyautogui.screenshot()
                img.show()
            if "shut down" in query:
                speak("Ok bye")
                os.system("shutdown /s /t 1")
            if "switch off" in query:
                speak("Ok bye")
                Mode = "OFF"








'''
Computer start greet

contact saving

youtube
'''